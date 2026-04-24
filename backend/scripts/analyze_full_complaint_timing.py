"""
全量投诉数据 — 投诉时间与发布时间间隔分布分析

直接从 ODPS 宽表拉取全量数据，只分析投诉时间，不跑大模型。
输出: 控制台统计 + PNG 分布图
"""
from __future__ import annotations

import json
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

# 将 backend/src 加入 sys.path 以便导入项目模块
_SRC = Path(__file__).resolve().parent.parent / "src"
sys.path.insert(0, str(_SRC))

from job_freshness.data_fetcher import fetch_freshness_candidates

# 招满关键词（含原始招满 + 联系不上）
_FILLED_KEYWORDS = [
    "已招满", "招满了", "人已找到", "人找到了", "已经找到",
    "不招了", "不招", "停招", "招到了", "已招到",
    "已经招到", "名额已满", "满了", "不需要了",
]
_UNREACHABLE_KEYWORDS = [
    "电话打不通", "联系不上", "没人接", "空号", "停机",
    "打不通", "无法接通", "关机", "不接电话",
]
_ALL_FILLED_KEYWORDS = _FILLED_KEYWORDS + _UNREACHABLE_KEYWORDS


def has_filled_complaint_signal(complaint_text: str) -> bool:
    """招满判断：招满关键词 + 联系不上关键词"""
    return any(kw in complaint_text for kw in _ALL_FILLED_KEYWORDS)

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib import rcParams

rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "DejaVu Sans"]
rcParams["axes.unicode_minus"] = False

OUT_DIR = Path(__file__).resolve().parent.parent / "output"

# 从投诉文本中提取所有投诉时间戳
_COMPLAINT_TS_RE = re.compile(r"【投诉\d+\s+([\d-]+\s+[\d:]+)】")


def parse_dt(s: str | None) -> datetime | None:
    if not s:
        return None
    try:
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is not None:
            dt = dt.replace(tzinfo=None)
        return dt
    except Exception:
        pass
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    return None


def extract_complaint_timestamps(text: str) -> list[datetime]:
    dts = []
    for m in _COMPLAINT_TS_RE.finditer(text):
        dt = parse_dt(m.group(1))
        if dt:
            dts.append(dt)
    return dts


BUCKET_ORDER = [
    "< 0h (异常)", "0-1h", "1-2h", "2-3h", "3-6h", "6-12h",
    "12-24h", "1-2天", "2-3天", "3-7天", "> 7天",
]


def hour_bucket(h: float) -> str:
    if h < 0: return "< 0h (异常)"
    if h < 1: return "0-1h"
    if h < 2: return "1-2h"
    if h < 3: return "2-3h"
    if h < 6: return "3-6h"
    if h < 12: return "6-12h"
    if h < 24: return "12-24h"
    if h < 48: return "1-2天"
    if h < 72: return "2-3天"
    if h < 168: return "3-7天"
    return "> 7天"


def fetch_data(bizdate: str) -> list[dict]:
    """从 ODPS 拉取全量数据（不过滤，拉回来再本地筛选）"""
    print(f"正在从 ODPS 拉取全量数据: bizdate={bizdate} ...")
    rows = fetch_freshness_candidates(bizdate, only_filled_complaints=False)
    print(f"拉取完成: {len(rows)} 条记录")
    return rows


def analyze_timing(rows: list[dict]) -> dict:
    """只分析招满投诉数据的投诉时间与发布时间间隔"""
    total = len(rows)
    has_complaint_rows = [r for r in rows if int(r.get("complaint_count") or 0) > 0]

    # 只保留含招满信号的投诉
    filled_rows = [r for r in has_complaint_rows
                   if has_filled_complaint_signal(str(r.get("complaint_content") or ""))]

    all_complaint_deltas = []       # 招满样本中所有投诉时间 - 发布时间
    first_complaint_deltas = []     # 招满样本第一条投诉 - 发布时间
    complaint_counts_per_job = []   # 每个招满职位的投诉条数

    # 按投诉条数分组
    complaint_count_buckets = Counter()

    for r in filled_rows:
        publish_time = parse_dt(str(r.get("publish_time") or ""))
        complaint_text = str(r.get("complaint_content") or "")
        complaint_dts = extract_complaint_timestamps(complaint_text)
        n_complaints = len(complaint_dts)
        complaint_counts_per_job.append(n_complaints)
        complaint_count_buckets[n_complaints] += 1

        if not publish_time or not complaint_dts:
            continue

        for cdt in complaint_dts:
            delta_h = (cdt - publish_time).total_seconds() / 3600
            all_complaint_deltas.append(delta_h)

        first_dt = min(complaint_dts)
        delta_h = (first_dt - publish_time).total_seconds() / 3600
        first_complaint_deltas.append(delta_h)

    return {
        "total": total,
        "has_complaint_count": len(has_complaint_rows),
        "filled_count": len(filled_rows),
        "all_complaint_deltas": all_complaint_deltas,
        "first_complaint_deltas": first_complaint_deltas,
        "complaint_counts_per_job": complaint_counts_per_job,
        "complaint_count_buckets": complaint_count_buckets,
    }


def print_report(data: dict):
    total = data["total"]
    print()
    print("=" * 80)
    print(f"  全量数据 · 招满投诉时间与发布时间间隔分析")
    print(f"  总职位数: {total} | 有投诉: {data['has_complaint_count']} | 招满投诉: {data['filled_count']}")
    print(f"  招满占有投诉比例: {data['filled_count']/max(data['has_complaint_count'],1)*100:.1f}%")
    print(f"  招满占全量比例: {data['filled_count']/max(total,1)*100:.1f}%")
    print("=" * 80)

    def print_stats(hours: list[float], label: str):
        if not hours:
            print(f"  {label}: 无数据")
            return
        s = sorted(hours)
        print(f"\n  {label} (n={len(hours)}):")
        print(f"    最小: {min(hours):.2f}h | 最大: {max(hours):.2f}h | 中位数: {s[len(s)//2]:.2f}h | 平均: {sum(hours)/len(hours):.2f}h")
        counter = Counter(hour_bucket(h) for h in hours)
        max_cnt = max(counter.values())
        for bucket in BUCKET_ORDER:
            cnt = counter.get(bucket, 0)
            if cnt == 0:
                continue
            bar = "█" * int(cnt / max_cnt * 40)
            print(f"    {bucket:14s} | {bar:40s} {cnt:5d} ({cnt/len(hours)*100:5.1f}%)")

    print("\n一、招满职位的投诉条数分布")
    print("-" * 60)
    cc = data["complaint_count_buckets"]
    for n in sorted(cc.keys()):
        print(f"  {n} 条投诉: {cc[n]:5d} 个职位 ({cc[n]/data['filled_count']*100:.1f}%)")

    print("\n二、招满投诉 — 所有投诉时间 vs 发布时间间隔（每条投诉单独计算）")
    print("-" * 60)
    print_stats(data["all_complaint_deltas"], "所有投诉")

    print("\n三、招满投诉 — 第一条投诉时间 vs 发布时间间隔（每个职位取最早投诉）")
    print("-" * 60)
    print_stats(data["first_complaint_deltas"], "第一条投诉")


def plot_charts(data: dict, bizdate: str):
    """生成招满投诉时间分布图"""
    fig, axes = plt.subplots(2, 2, figsize=(18, 14))
    fig.suptitle(
        f"全量数据 · 招满投诉时间与发布时间间隔分布\n"
        f"bizdate={bizdate} | 总职位={data['total']} | 有投诉={data['has_complaint_count']} | 招满={data['filled_count']}",
        fontsize=15, fontweight="bold", y=0.98
    )
    plt.subplots_adjust(hspace=0.35, wspace=0.28, top=0.90, bottom=0.06)

    C_BLUE = "#4C72B0"
    C_ORANGE = "#DD8452"
    C_GREEN = "#55A868"

    # ① 所有投诉时间 vs 发布时间（招满样本）
    ax = axes[0, 0]
    hrs = [h for h in data["all_complaint_deltas"] if -30 <= h <= 200]
    if hrs:
        ax.hist(hrs, bins=80, color=C_ORANGE, edgecolor="white", alpha=0.85)
        median_v = sorted(data["all_complaint_deltas"])[len(data["all_complaint_deltas"]) // 2]
        ax.axvline(median_v, color="red", linestyle="--", linewidth=1.5, label=f"中位数 {median_v:.1f}h")
        ax.legend(fontsize=10)
    ax.set_xlabel("间隔（小时）", fontsize=11)
    ax.set_ylabel("投诉条数", fontsize=11)
    ax.set_title(f"① 招满投诉 — 所有投诉时间 vs 发布时间\n（每条投诉单独计算, n={len(data['all_complaint_deltas'])}）", fontsize=12)
    ax.set_xlim(-30, 200)

    # ② 第一条投诉 vs 发布时间（招满样本）
    ax = axes[0, 1]
    hrs = [h for h in data["first_complaint_deltas"] if -30 <= h <= 200]
    if hrs:
        ax.hist(hrs, bins=80, color=C_GREEN, edgecolor="white", alpha=0.85)
        median_v = sorted(data["first_complaint_deltas"])[len(data["first_complaint_deltas"]) // 2]
        ax.axvline(median_v, color="red", linestyle="--", linewidth=1.5, label=f"中位数 {median_v:.1f}h")
        ax.legend(fontsize=10)
    ax.set_xlabel("间隔（小时）", fontsize=11)
    ax.set_ylabel("职位数", fontsize=11)
    ax.set_title(f"② 招满投诉 — 第一条投诉 vs 发布时间\n（每个职位取最早投诉, n={len(data['first_complaint_deltas'])}）", fontsize=12)
    ax.set_xlim(-30, 200)

    # ③ 0-24h 细粒度分布（第一条投诉，招满样本）
    ax = axes[1, 0]
    hrs_24 = [h for h in data["first_complaint_deltas"] if 0 <= h <= 24]
    if hrs_24:
        ax.hist(hrs_24, bins=48, color=C_GREEN, edgecolor="white", alpha=0.85)
        median_v = sorted(hrs_24)[len(hrs_24) // 2]
        ax.axvline(median_v, color="red", linestyle="--", linewidth=1.5, label=f"中位数 {median_v:.1f}h")
        ax.legend(fontsize=10)
    ax.set_xlabel("间隔（小时）", fontsize=11)
    ax.set_ylabel("职位数", fontsize=11)
    ax.set_title(f"③ 招满投诉 — 第一条投诉 0-24h 细粒度\n（n={len(hrs_24)}）", fontsize=12)

    # ④ 分段柱状图
    ax = axes[1, 1]
    bucket_labels = ["<0h", "<1h", "1-2h", "2-3h", "3-6h", "6-12h", "12-24h", "1-2天", "2-3天", "3-7天", ">7天"]
    bucket_ranges = [(-9999, 0), (0, 1), (1, 2), (2, 3), (3, 6), (6, 12), (12, 24), (24, 48), (48, 72), (72, 168), (168, 99999)]

    def count_buckets(hours):
        return [sum(1 for h in hours if lo <= h < hi) for lo, hi in bucket_ranges]

    all_counts = count_buckets(data["all_complaint_deltas"]) if data["all_complaint_deltas"] else [0] * len(bucket_labels)
    first_counts = count_buckets(data["first_complaint_deltas"]) if data["first_complaint_deltas"] else [0] * len(bucket_labels)

    x = range(len(bucket_labels))
    width = 0.35
    bars1 = ax.bar([i - width/2 for i in x], first_counts, width, label="第一条投诉", color=C_GREEN, alpha=0.85)
    bars2 = ax.bar([i + width/2 for i in x], all_counts, width, label="所有投诉", color=C_ORANGE, alpha=0.85)
    ax.set_xticks(list(x))
    ax.set_xticklabels(bucket_labels, fontsize=9, rotation=30)
    ax.set_ylabel("数量", fontsize=11)
    ax.set_title("④ 招满投诉 — 时间间隔分段统计", fontsize=12)
    ax.legend(fontsize=10)
    for bar in bars1:
        h = bar.get_height()
        if h > 0:
            ax.text(bar.get_x() + bar.get_width()/2, h + 0.3, str(int(h)), ha="center", va="bottom", fontsize=7)
    for bar in bars2:
        h = bar.get_height()
        if h > 0:
            ax.text(bar.get_x() + bar.get_width()/2, h + 0.3, str(int(h)), ha="center", va="bottom", fontsize=7)

    out_path = OUT_DIR / f"filled_complaint_timing_{bizdate}.png"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(str(out_path), dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"\n图表已保存: {out_path}")


def export_excel(data: dict, bizdate: str):
    """将分析结果写入 Excel 表格"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

    def style_data(ws, start_row, end_row):
        for r in range(start_row, end_row + 1):
            for cell in ws[r]:
                cell.border = thin_border
                cell.alignment = center

    # ── Sheet 1: 概览 ──
    ws = wb.active
    ws.title = "概览"
    ws.append(["指标", "数值"])
    style_header(ws)
    overview = [
        ("数据日期", bizdate),
        ("总职位数", data["total"]),
        ("有投诉职位数", data["has_complaint_count"]),
        ("招满投诉职位数", data["filled_count"]),
        ("招满占有投诉比例", f"{data['filled_count']/max(data['has_complaint_count'],1)*100:.1f}%"),
        ("招满占全量比例", f"{data['filled_count']/max(data['total'],1)*100:.1f}%"),
        ("", ""),
        ("招满关键词", "已招满/招满了/人已找到/人找到了/已经找到/不招了/不招/停招/招到了/已招到/已经招到/名额已满/满了/不需要了"),
        ("联系不上关键词", "电话打不通/联系不上/没人接/空号/停机/打不通/无法接通/关机/不接电话"),
    ]
    for label, val in overview:
        ws.append([label, val])
    style_data(ws, 2, ws.max_row)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80

    # ── Sheet 2: 投诉条数分布 ──
    ws2 = wb.create_sheet("投诉条数分布")
    ws2.append(["投诉条数", "职位数", "占比"])
    style_header(ws2)
    cc = data["complaint_count_buckets"]
    for n in sorted(cc.keys()):
        ws2.append([n, cc[n], f"{cc[n]/data['filled_count']*100:.1f}%"])
    style_data(ws2, 2, ws2.max_row)
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 10

    # ── Sheet 3: 时间间隔分段统计 ──
    ws3 = wb.create_sheet("时间间隔分段统计")
    ws3.append(["时间段", "所有投诉(条数)", "所有投诉(占比)", "第一条投诉(职位数)", "第一条投诉(占比)"])
    style_header(ws3)

    bucket_labels = ["< 0h (异常)", "0-1h", "1-2h", "2-3h", "3-6h", "6-12h", "12-24h", "1-2天", "2-3天", "3-7天", "> 7天"]
    bucket_ranges = [(-9999, 0), (0, 1), (1, 2), (2, 3), (3, 6), (6, 12), (12, 24), (24, 48), (48, 72), (72, 168), (168, 99999)]

    all_hrs = data["all_complaint_deltas"]
    first_hrs = data["first_complaint_deltas"]
    n_all = max(len(all_hrs), 1)
    n_first = max(len(first_hrs), 1)

    for label, (lo, hi) in zip(bucket_labels, bucket_ranges):
        cnt_all = sum(1 for h in all_hrs if lo <= h < hi)
        cnt_first = sum(1 for h in first_hrs if lo <= h < hi)
        ws3.append([
            label,
            cnt_all, f"{cnt_all/n_all*100:.1f}%",
            cnt_first, f"{cnt_first/n_first*100:.1f}%",
        ])
    # 合计行
    ws3.append(["合计", len(all_hrs), "100%", len(first_hrs), "100%"])
    style_data(ws3, 2, ws3.max_row)
    for col in ["A", "B", "C", "D", "E"]:
        ws3.column_dimensions[col].width = 18

    # ── Sheet 4: 统计摘要 ──
    ws4 = wb.create_sheet("统计摘要")
    ws4.append(["统计项", "所有投诉(每条)", "第一条投诉(每职位)"])
    style_header(ws4)

    def safe_stats(hours):
        if not hours:
            return {"min": "-", "max": "-", "median": "-", "mean": "-"}
        s = sorted(hours)
        return {
            "min": f"{min(hours):.2f}h",
            "max": f"{max(hours):.2f}h",
            "median": f"{s[len(s)//2]:.2f}h",
            "mean": f"{sum(hours)/len(hours):.2f}h",
        }

    s_all = safe_stats(all_hrs)
    s_first = safe_stats(first_hrs)
    for key, label in [("min", "最小值"), ("max", "最大值"), ("median", "中位数"), ("mean", "平均值")]:
        ws4.append([label, s_all[key], s_first[key]])
    ws4.append(["样本数", len(all_hrs), len(first_hrs)])
    style_data(ws4, 2, ws4.max_row)
    ws4.column_dimensions["A"].width = 16
    ws4.column_dimensions["B"].width = 22
    ws4.column_dimensions["C"].width = 22

    # 保存
    out_path = OUT_DIR / f"filled_complaint_timing_{bizdate}.xlsx"
    wb.save(str(out_path))
    print(f"Excel 已保存: {out_path}")


def main():
    bizdate = sys.argv[1] if len(sys.argv) > 1 else "20260420"
    print(f"分析日期: {bizdate}")

    # 尝试从本地缓存读取
    cache_path = OUT_DIR / f"full_candidates_{bizdate}.json"
    if cache_path.exists():
        print(f"从本地缓存读取: {cache_path}")
        rows = json.loads(cache_path.read_text(encoding="utf-8"))
    else:
        rows = fetch_data(bizdate)
        # 缓存到本地
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        # 转换 ODPS 返回的特殊类型
        serializable_rows = []
        for r in rows:
            sr = {}
            for k, v in r.items():
                sr[k] = str(v) if v is not None else None
            serializable_rows.append(sr)
        cache_path.write_text(json.dumps(serializable_rows, ensure_ascii=False), encoding="utf-8")
        print(f"数据已缓存: {cache_path}")
        rows = serializable_rows

    data = analyze_timing(rows)
    print_report(data)
    plot_charts(data, bizdate)
    export_excel(data, bizdate)


if __name__ == "__main__":
    main()
